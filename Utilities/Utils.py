import inspect
import logging

import softest
from openpyxl import Workbook, load_workbook


class Utils(softest.TestCase):

    def assert_list_item_text(self, list_items, value):
        for stop in list_items:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("Test Passed" + stop.text)
            else:
                print("Test Failed")
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):

        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        # create console handler and set level to debug
        ch = logging.StreamHandler()
        ch.setLevel(logLevel)

        # ALSO USE FILE HANDLER
        fh = logging.FileHandler("automation.log")

        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        formatter1 = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                                       datefmt='%m/%d/%Y %I:%M:%S %p')

        # add formatter to ch
        ch.setFormatter(formatter)
        fh.setFormatter(formatter1)

        # add ch to logger
        logger.addHandler(ch)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        max_row = sh.max_row
        max_col = sh.max_column
        for i in range(2, max_row + 1):
            row = []
            for j in range(1, max_col + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)

        return data_list
